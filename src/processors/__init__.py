from pydantic import ValidationError as PydanticValidationError
from rapidfuzz import process,fuzz
from typing import Any
from excel import aditional_data
from excel.aditional_data import AditionalDataManager
from models.lineup import LineUpBaseModel
from pathlib import Path

import excel
import openpyxl
import constants
import logging
import validations
import reports

logger = logging.getLogger()

def row_to_dict( row : tuple, layout : type[excel.LineUpLayout] | type[excel.LineUpLayoutVariant])->dict[str, Any]:
   return {
      member.name: (row[member.col - 1] if member.col - 1 < len(row) else None)
      for member in layout
   }


def field_to_col(field_name: str, layout) -> int:
   """
   Mapea el nombre del campo Pydantic (member.name del enum, ej. DATE_OF_ARRIVAL)
   al número de columna Excel del layout. Retorna -1 si no se encuentra.
   """
   for member in layout:
      if member.name == field_name:
         return member.col
   return -1

def aditional_validations(lineup_model : LineUpBaseModel, additional_data : AditionalDataManager| None):
   if aditional_data is None:
      return None
      
def depurate_lineup_files(folder_path : Path, offices_key : list[str], min_score : int, header_row : int, additional : AditionalDataManager | None , compare_aditional : bool):
   """
   Funcion para cargar los archivos de excel de las oficinas que esten en una ruta y develve las hojas esperadas para cada una 
   """

   xlsx_files = list(folder_path.glob('*.xlsx'))
   xlsx_files_names = [f.name.upper() for f in xlsx_files]

   if not xlsx_files:
      logger.warning("No se encontraron archivos .xlsx en '%s'.", folder_path)

   if len(xlsx_files) != len(offices_key):
      logger.warning(
         "No hay los archivos suficientes para procesar todas las oficinas"
      )

   checker = validations.OverlapChecker()     
   depuration_result = validations.DepurationResult()
      
   for office in offices_key:
      result = process.extractOne(office,xlsx_files_names, scorer=fuzz.partial_ratio)
      if result is None:
         logger.warning(
            "No se encontró ninguna hoja para '%s'. Hojas disponibles: %s",
            office,
            xlsx_files_names,
         )
         continue
      office_match,score, idx = result
      if score < min_score:
         logger.warning(
            "Mejor coincidencia para '%s' fue '%s' con score %d (mínimo requerido: %d). "
            "Se retorna None.",
            office,
            office_match,
            score,
            min_score,
         )
         continue
      file_path = xlsx_files[idx]
      xlsx_files.pop(idx) ; xlsx_files_names.pop(idx)

      logger.info(f'Procesando archivo de la oficina %s con nombre %s ',office,file_path.name)
      file_expected_sheets = constants.EXPECTED_SHEETS.get(office)
      assert file_expected_sheets is not None,f"No existe configuración de hojas para la oficina '{office}'. " f"Claves disponibles: {list(constants.EXPECTED_SHEETS.keys())}"

      wb = openpyxl.load_workbook(file_path,data_only=True)

      available_sheets = wb.sheetnames
      logger.info(
         "Archivo '%s' cargado, Hojas disponibles: %s ",
         file_path.name,available_sheets
      )

      office_result = validations.OfficeResult(office,file_path=file_path)
      assert additional
      available_products = additional.get_available_products()
      for expected_name in file_expected_sheets:
         matched = excel.find_best_sheet_match(available_sheets,expected_name, min_score)

         if matched is None:
            logger.warning("Hoja '%s' de oficina '%s' es None, se omite.", expected_name, office)
            continue

         available_sheets.remove(matched)
                  
         ws = wb[matched]
         
         layout = constants.PORT_LAYOUTS.get(expected_name, excel.LineUpLayout )
         max_col = max(member.col for member in layout)
         
         # Antes de iterar, vamos a mirar si la primera fila (HEADER_ROW, col = 2), corresponde con buena concidencia > 90% al primer valor de layout
         headers_ok = excel.verify_headers(ws,layout, header_row)
         if not headers_ok:
            logger.warning(
               "Hoja '%s' de oficina '%s' tiene headers que no coinciden con layout '%s'. "
               "Se continúa procesando, pero los datos pueden ser incorrectos.",
               matched, office, layout.__name__,
             )
         error_registry = validations.ErrorRegistry()
         sheet_rows = []
         available_terminals = additional.get_available_terminals(expected_name)
         
         for excel_row, row in enumerate(ws.iter_rows(min_row=header_row+1,max_col=max_col,values_only=True), start=header_row+1 ):
            if all(cell is None for cell in row):
               break
            
            row_dict = row_to_dict(row, layout)
            if layout == excel.LineUpLayout:
               row_dict['WINDOWS'] = None

            try:
               validated = LineUpBaseModel.model_validate(row_dict, context={'available_terminals':available_terminals, 'available_products':available_products})
               sheet_rows.append(validated.to_client_report())

               interval = validations.build_interval(excel_row, validated,expected_name)
               if interval is not None:
                  checker.add(interval)

            except PydanticValidationError as exc:
               for error in exc.errors():
                  loc = error["loc"]
                  
                  ctx = error.get("ctx", {})

                  # field validator → loc tiene el campo; model validator → loc vacío pero ctx tiene 'fields'
                  if loc:
                      fields_involved = [str(loc[0])]
                  else:
                      fields_involved = ctx.get("fields", ["MODEL"])

                  msg_template = error["msg"]

                  field_values = ctx.get("values", {})
                  try:
                      msg = msg_template.format(**field_values)
                      print(msg)
                  except (KeyError, ValueError):
                      msg = msg_template

                  for field_name in fields_involved:
                      col = field_to_col(field_name, layout)
                      val = field_values.get(field_name, error.get("input"))

                      logger.warning(
                          "Fila %d, col %d, campo '%s': %s",
                          excel_row, col, field_name, msg,
                      )
                      error_registry.add(
                           row=excel_row, col=col, field=field_name,
                           msg=msg, value=val, source="pydantic"
                     )
         office_result.sheets.append(
            validations.SheetResult(
               expected_name=expected_name,
               matched_name=matched,
               headers_ok=headers_ok,
               rows=sheet_rows,
               registry=error_registry,
            )
         )
      depuration_result.offices.append(office_result)
      wb.close()

   conflicts = checker.check()
   depuration_result.overlaps = conflicts
   for conflict in conflicts:
       logger.warning(
           "Solapamiento: VESSEL '%s' fila %d vs fila %d.",
           conflict.vessel, conflict.row_a, conflict.row_b
       )

   return depuration_result 

__all__ = ['depurate_lineup_files']
