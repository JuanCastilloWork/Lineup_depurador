
from datetime import date
from decimal import Decimal
from typing import TYPE_CHECKING
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
import io
from pathlib import Path
from excel import layots
from processors.final_processor import PortBundle
import validations
import pandas as pd

HEADER_BG = 'FFFFFF'
HEADER_FG = '000000'
HEADER_TITLE_BG = '122649'
HEADER_TITLE_FG = 'BFBFBF'
BORDER_COLOR = '000000'

COL_HEADER_BG = '122649'
COL_HEADER_FG = 'BFBFBF'
ROW_ALT_FG = '002060'

ERROR_FILL_COLOR = 'FF4C4C'   # rojo para celdas con error

_thin = Side(style='thin', color = BORDER_COLOR)
_dashed = Side(style='dashed',color = BORDER_COLOR)
_cell_border_thin = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_cell_border_dashed = Border(left=_dashed, right=_dashed, top=_dashed, bottom=_dashed)

def _apply_border_to_merged_range(ws: Worksheet, cell_range: str, border: Border):
   """Aplica border a todas las celdas de un rango mergeado."""
   from openpyxl.utils import range_boundaries
   min_col, min_row, max_col, max_row = range_boundaries(cell_range)
   for row in range(min_row, max_row + 1):
       for col in range(min_col, max_col + 1):
           ws.cell(row=row, column=col).border = border

def _style_col_header(cell):

   cell.font      = Font(name="Calibri", color=COL_HEADER_FG, size=11)
   cell.fill      = PatternFill("solid", start_color=COL_HEADER_BG)
   cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
   cell.border    = _cell_border_dashed


def _style_data_cell(cell: Cell, alternate: bool = False, has_error: bool = False):
    fg = ROW_ALT_FG if alternate else "000000"
    cell.font      = Font(name="Calibri", size=9, color=fg, bold=alternate)
    cell.fill      = (
        PatternFill("solid", start_color=ERROR_FILL_COLOR)
        if has_error
        else PatternFill("solid", start_color='FFFFFF')
    )
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    cell.border    = _cell_border_thin

def _make_anchor(ws, col_letter, row_num, offset_x_px, offset_y_px, img_w_px, img_h_px):
   """Crea un OneCellAnchor con offsets en EMU para centrar la imagen."""
   from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
   from openpyxl.utils import column_index_from_string
   
   col_idx = column_index_from_string(col_letter) - 1  # 0-indexed
   row_idx = row_num - 1  # 0-indexed
   
   marker = AnchorMarker(
       col=col_idx,
       colOff=pixels_to_EMU(offset_x_px),
       row=row_idx,
       rowOff=pixels_to_EMU(offset_y_px),
   )
   anchor = OneCellAnchor(_from=marker, ext=None)
   anchor.ext.cx = pixels_to_EMU(img_w_px)
   anchor.ext.cy = pixels_to_EMU(img_h_px)
   return anchor

    
class LineUpExcelReport:
   
   HEADER_ROW = 12
   
   def __init__(self ,
      bundle : dict[str,PortBundle],      
      company: str,
      lineup_date: date,
                ) -> None:

      self._bundle = bundle
      self._company             = company.upper()
      self._lineup_date         = lineup_date
      self._current_cod         = 'DB-GOP - FT - 16 / 3'
      self._current_cod_date    = date(2023, 9, 8)


   def _build_error_index(
        self, report: validations.ValidationReport | None
    ) -> dict[str, set[int]]:
      """
      Devuelve {attr_name: {row_indices con error}}.
      row_index en CellError corresponde al índice posicional del DataFrame (0-based).
      """
      if report is None:
          return {}
 
      index: dict[str, set[int]] = {}
      for error in report.all_errors:
         index.setdefault(error.column, set()).add(error.row_index)
      return index

   @staticmethod
   def _layout_attr_map(layout_cls: type[layots.LineUpBaseLayout]) -> dict[str, object]:
      """
      Devuelve {attr_name: ColDef} para todos los ColDef del layout.
      Ejemplo: {"DATE_OF_ARRIVAL": ColDef(col=3, name="DATE OF ARRIVAL", ...)}
      """
      result = {}
      for member in layout_cls:          # itera miembros del Enum
          val = member.value             # <-- el ColDef real
          if hasattr(val, 'col') and hasattr(val, 'name'):
              result[member.name] = val  # member.name = "DATE_OF_ARRIVAL", val.name = "DATE OF ARRIVAL"
      return result
      
   def _write_data_columns(
      self,
      ws: Worksheet,
      df: pd.DataFrame,
      layout_cls: type[layots.LineUpBaseLayout],
      error_index: dict[str, set[int]],
      header_row: int,
      agency_col_name: str = "AGENCY",   # nombre del ColDef attr para AGENCY
    ):
      """
      Itera por ColDef (columna a columna).  Para cada ColDef busca la Serie
      del DataFrame cuyo nombre de columna coincide con ColDef.name y escribe
      todas las celdas de esa columna de una vez.
      """
      attr_map = self._layout_attr_map(layout_cls)
 
      # Precalculamos filas que pertenecen a la compañía (deep_blue)
      # Buscamos el ColDef cuyo attr_name sea AGENCY
      agency_col_def = attr_map.get(agency_col_name)
      deep_blue_rows: set[int] = set()
      if agency_col_def is not None and agency_col_name in df.columns:
         deep_blue_rows = set(
            df.index[df[agency_col_def.name] == self._company].tolist()
         )
 
      for attr_name, col_def in attr_map.items():
         # La serie del DataFrame se identifica por col_def.name
         if attr_name not in df.columns:
            continue
 
         series: pd.Series = df[attr_name]
         error_rows = error_index.get(attr_name, set())
 
         for positional_idx, value in enumerate(series):
            excel_row = header_row + 1 + positional_idx
            cell = ws.cell(row=excel_row, column=col_def.col)
 
            has_error  = positional_idx in error_rows
            is_alt     = positional_idx in deep_blue_rows
 
            # Valor
            cell.value = value
 
            # Estilo base
            _style_data_cell(cell, alternate=is_alt, has_error=has_error)
 
            # Alineaciones especiales por columna
            if attr_name == 'PIER':
               cell.alignment = Alignment(horizontal='center')
            elif attr_name in ('MT_BY_PRODUCT', 'TOTAL_MT'):
               cell.alignment = Alignment(horizontal='right')
               
               if isinstance(value, Decimal):
                  if value % 1 == Decimal('0'):
                     cell.number_format = "#,##0"
                  else:
                     cell.number_format = "#,##0.##"
               

   def _cm_to_charunit_aprox(self, value : float):
      return value


   def _change_column_widths(self, ws : Worksheet, layout_cls : type[layots.LineUpBaseLayout]):
      
      ws.column_dimensions['A'].width = self._cm_to_charunit_aprox(1.75)
      ws.column_dimensions['B'].width = self._cm_to_charunit_aprox(22.67)
      ws.column_dimensions['C'].width = self._cm_to_charunit_aprox(15)
      ws.column_dimensions['D'].width = self._cm_to_charunit_aprox(14)
      ws.column_dimensions['E'].width = self._cm_to_charunit_aprox(8)
      ws.column_dimensions['F'].width = self._cm_to_charunit_aprox(14)
      ws.column_dimensions['G'].width = self._cm_to_charunit_aprox(19)
      ws.column_dimensions['H'].width = self._cm_to_charunit_aprox(11.85)
      ws.column_dimensions['I'].width = self._cm_to_charunit_aprox(16.1)
      ws.column_dimensions['J'].width = self._cm_to_charunit_aprox(21)
      ws.column_dimensions['K'].width = self._cm_to_charunit_aprox(21)
      ws.column_dimensions['L'].width = self._cm_to_charunit_aprox(12.6)
      ws.column_dimensions['M'].width = self._cm_to_charunit_aprox(19)
      ws.column_dimensions['N'].width = self._cm_to_charunit_aprox(28)
      ws.column_dimensions['O'].width = self._cm_to_charunit_aprox(34)
      ws.column_dimensions['P'].width = self._cm_to_charunit_aprox(16)
      ws.column_dimensions['Q'].width = self._cm_to_charunit_aprox(32)
      if layout_cls == layots.LineUpReportVariantLayout:
         ws.column_dimensions['R'].width = 24
   def _write_header_block(self, ws : Worksheet,layout_cls : type[layots.LineUpBaseLayout], port : str, logo_path : Path ):
      ws.merge_cells(f'B2:D7')
      ws.merge_cells(f'E2:M7')
      ws.merge_cells(f'N2:O7')
      if layout_cls == layots.LineUpReportVariantLayout:
         
         ws.merge_cells(f'P2:R7')
      else:
         ws.merge_cells(f'P2:Q7')

      # Logo va en la B2:D7
      PADDING_PX = 6       
    
      # Calcular altura total del bloque B2:D7 en puntos (filas 2 a 7)

      total_height_pt = sum(
          ws.row_dimensions[r].height or 15  # 15pt es el default de openpyxl
          for r in range(2, 8)
      )   
      # Calcular ancho total de columnas B, C, D en píxeles
      # openpyxl: 1 unidad de ancho de col ≈ 7px (aprox)
      COL_UNIT_TO_PX = 7
      col_letters = ['B', 'C', 'D']

      
      total_width_px = sum(
          (ws.column_dimensions[c].width or 8) * COL_UNIT_TO_PX
          for c in col_letters
      )
      # Altura disponible para la imagen (puntos → píxeles, 1pt ≈ 1.333px)
      PT_TO_PX = 1.333
      available_height_px = int(total_height_pt * PT_TO_PX) - (PADDING_PX * 2)

      # Redimensionar logo con Pillow manteniendo aspect ratio
      with PILImage.open(logo_path) as img:
          orig_w, orig_h = img.size
          scale = min(total_width_px / orig_w, available_height_px / orig_h)
          new_w = int(orig_w * scale)
          new_h = int(orig_h * scale)
          img_resized = img.resize((new_w, new_h), PILImage.Resampling.LANCZOS)
          
          buf = io.BytesIO()
          fmt = logo_path.suffix.lstrip('.').upper()
          fmt = 'PNG' if fmt not in ('PNG', 'JPEG', 'JPG') else fmt
          img_resized.save(buf, format='PNG')
          buf.seek(0)

      xl_img = XLImage(buf)
      xl_img.width  = new_w
      xl_img.height = new_h

      # Calcular offset para centrar horizontalmente y con padding vertical
      offset_x = int((total_width_px - new_w) / 2)
      offset_y = PADDING_PX

      xl_img.anchor = 'B2'
      xl_img.anchor = _make_anchor(ws, 'B', 2, offset_x, offset_y, new_w, new_h)
      ws.add_image(xl_img)

      # Titulo
      title_cell : Cell = ws['E2']
      title_cell.value = f'{port.upper()} PORT LINE - UP'
      title_cell.font      = Font(name="Calibri", bold=True, size=14, color=HEADER_TITLE_FG)
      title_cell.fill      = PatternFill("solid", start_color=HEADER_TITLE_BG)
      title_cell.alignment = Alignment(horizontal="center", vertical="center")

      
      # Cod
      cod_cell : Cell = ws['N2']
      cod_cell.value = f'COD: {self._current_cod.upper()}'
      cod_cell.font      = Font(name="Calibri", bold=True, size=11, color=HEADER_FG)
      cod_cell.fill      = PatternFill("solid", start_color=HEADER_BG)
      cod_cell.alignment = Alignment(horizontal="center", vertical="center")

      # Cod
      cod_cell : Cell = ws['P2']
      cod_cell.value = f'DATE: {self._current_cod_date.strftime('%d-%m-%Y')}'
      cod_cell.font      = Font(name="Calibri", bold=True, size=11, color=HEADER_FG)
      cod_cell.fill      = PatternFill("solid", start_color=HEADER_BG)
      cod_cell.alignment = Alignment(horizontal="center", vertical="center")

      _apply_border_to_merged_range(ws, 'N2:O7', _cell_border_thin)
      if layout_cls == layots.LineUpReportVariantLayout:
         _apply_border_to_merged_range(ws, 'P2:R7', _cell_border_thin)
      else:
         _apply_border_to_merged_range(ws, 'P2:Q7', _cell_border_thin)
      
      # Supongo que pongamos la fecha como header XD
      # Falta centrar la fecha, y a la b10 ponerla en negrilla, y que este en size 11 vs 12 el valor

      cell = ws['B10']
      cell.value = 'Date:'
      cell.font      = Font(name="Calibri", bold=True, size=12)
      cell.alignment = Alignment(horizontal="center", vertical="center")
      cell = ws['C10']
      cell.value = self._lineup_date.strftime('%d-%m-%Y')      
      cell.font      = Font(name="Calibri", size=11)
      cell.alignment = Alignment(horizontal="center", vertical="center")

   def create_report(self, output_path : Path = Path('daily_lineup.xlsx'), header_row : int = 12, logo_path : Path = Path('./assets/company_logo.png')):

      wb = Workbook()
      wb.remove(wb.active)
 
      for port_name, port_bundle in self._bundle.items():
         layout_cls  = port_bundle.layout
         report      = port_bundle.report
         error_index = self._build_error_index(report)
         sheet_name = port_name.upper().replace('_',' ')   
         sheet: Worksheet = wb.create_sheet(sheet_name)
         sheet.sheet_view.showGridLines = False
 
         self._change_column_widths(sheet, layout_cls)
         self._write_header_block(sheet, layout_cls, port_name, logo_path)
 
         # Headers de columna
         attr_map = self._layout_attr_map(layout_cls)
         for col_def in attr_map.values():
             cell = sheet.cell(header_row, col_def.col)
             cell.value = col_def.name      # label visible en Excel = ColDef.name
             _style_col_header(cell)
 
         # Datos (iteración por columna)
         self._write_data_columns(
             ws=sheet,
             df=port_bundle.df,
             layout_cls=layout_cls,
             error_index=error_index,
             header_row=header_row,
         )
 
      wb.save(output_path)
      return output_path
