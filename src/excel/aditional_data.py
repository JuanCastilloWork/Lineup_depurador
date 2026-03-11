from openpyxl import load_workbook
import utils
from constants import TypeEnum
from pathlib import Path


class AditionalDataManager:
   def __init__(self, path: Path, eager: bool = False) -> None:

      self.path = path
      self.product_types: dict[TypeEnum,list[str]] = {}
      self.port_terminals: dict[str, list[str]] = {}
      self.companies: dict[str, tuple[bool, bool, bool]] = {}
      self.charterers: list[str] = []
      self.shipowners: list[str] = []
      self.agency: list[str] = []
      if eager:
         self.load()
      self.first_load()

   def get_charterers(self):
      return self.charterers

   def get_available_products(self):
      return self.product_types

   def get_available_terminals(self, port : str):
      return self.port_terminals.get(port,[])
   
   def get_shipowners(self):
      return self.shipowners

   def get_agencies(self):
      return self.agency

   def first_load(self):
      wb = load_workbook(self.path, data_only=True)
      try:
         self._load_products_types(wb)
         self._load_port_terminals(wb)
      finally:
         wb.close()

   def load(self):
      wb = load_workbook(self.path, data_only=True)
      try:
         self._load_products_types(wb)
         self._load_companies(wb)
         self._load_port_terminals(wb)
      finally:
         wb.close()

   def _load_products_types(self, wb):
      ws = wb['PRODUCT']
      table = ws.tables['PRODUCTS']
      data = ws[table.ref]
      headers = [cell.value for cell in data[0]]
      try:
         idx_product = headers.index('PRODUCT')
         idx_type    = headers.index('TYPE')
      except ValueError as e:
         raise ValueError(f"Columna no encontrada en tabla PRODUCTS: {e}")
      for row in data[1:]:
         raw_product = row[idx_product].value
         raw_type    = row[idx_type].value
         product  = utils.remove_multiple_white_spaces(raw_product)
         type_str = utils.remove_multiple_white_spaces(raw_type)
         if product is None or type_str is None:
            break
         product  = product.upper()
         type_str = type_str.upper()

         try:
            type_enum = TypeEnum(type_str)
            if type_enum not in self.product_types:
               self.product_types[type_enum] = []
            self.product_types[type_enum].append(product)
         except ValueError:
            raise ValueError(f"Tipo no reconocido en Excel: '{type_str}' para producto '{product}'")

   def _load_companies(self, wb):
      ws = wb['COMPANIES']
      table = ws.tables['COMPANIES']
      data = ws[table.ref]
      headers = [cell.value for cell in data[0]]
      try:
         idx_name        = headers.index('NAME')
         idx_shipowner   = headers.index('IS_SHIPOWNER')
         idx_charterer   = headers.index('IS_CHARTERER')
         idx_agency      = headers.index('IS_AGENCY')
      except ValueError as e:
         raise ValueError(f"Columna no encontrada en tabla COMPANIES: {e}")

      for row in data[1:]:
         raw_name = row[idx_name].value
         name = utils.remove_multiple_white_spaces(raw_name)
         if name is None:
            break
         name = name.upper()

         is_shipowner = bool(row[idx_shipowner].value)
         is_charterer = bool(row[idx_charterer].value)
         is_agency    = bool(row[idx_agency].value)

         # Si ya existe, limpiar de las listas para reemplazar
         if name in self.companies:
            old_ship, old_chart, old_agcy = self.companies[name]
            if old_ship  and name in self.shipowners:  self.shipowners.remove(name)
            if old_chart and name in self.charterers:  self.charterers.remove(name)
            if old_agcy  and name in self.agency:      self.agency.remove(name)

         self.companies[name] = (is_shipowner, is_charterer, is_agency)

         if is_shipowner: self.shipowners.append(name)
         if is_charterer: self.charterers.append(name)
         if is_agency:    self.agency.append(name)

   def _load_port_terminals(self, wb):
      ws = wb['PORT']
      table = ws.tables['TERMINALS']
      data = ws[table.ref]
      headers = [cell.value for cell in data[0]]
      try:
         idx_port     = headers.index('PORT')
         idx_terminal = headers.index('TERMINAL')
      except ValueError as e:
         raise ValueError(f"Columna no encontrada en tabla TERMINALS: {e}")

      for row in data[1:]:
         raw_port     = row[idx_port].value
         raw_terminal = row[idx_terminal].value
         port     = utils.remove_multiple_white_spaces(raw_port)
         terminal = utils.remove_multiple_white_spaces(raw_terminal)
         if port is None:
            break
         port = port.upper()
         terminal = terminal.upper() if terminal else None

         if port not in self.port_terminals:
            self.port_terminals[port] = []

         if terminal and terminal not in self.port_terminals[port]:
            self.port_terminals[port].append(terminal)
