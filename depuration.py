
import openpyxl
from pathlib import Path
from openpyxl.worksheet.worksheet import Worksheet
from rapidfuzz import process,fuzz
from config import *
from typing import Any
from error_registry import CellError, ErrorRegistry
from models import LineUpBaseModel
from pydantic import ValidationError as PydanticValidationError
from overlap import OverlapChecker, OverlapConflict, build_interval
from results import DepurationResult, OfficeResult, SheetResult
import logging
import sys
 
file_handler = logging.FileHandler('depuration.log')
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)

logger = logging.getLogger()
logger.addHandler(file_handler)
logger.setLevel(logging.INFO)

def find_best_sheet_match(sheet_names : list[str], expected_name : str,min_score : int = MIN_MATCH_SCORE)->str | None:
   result = process.extractOne(expected_name,sheet_names, scorer= fuzz.WRatio)
   if result is None:
      logger.warning(
         "No se encontró ninguna hoja para '%s'. Hojas disponibles: %s",
         expected_name,
         sheet_names,
      )
      return None
   matched_name,score,_ = result
   if score < min_score:
      logger.warning(
         "Mejor coincidencia para '%s' fue '%s' con score %d (mínimo requerido: %d). "
         "Se retorna None.",
         expected_name,
         matched_name,
         score,
         min_score,
      )
      return None
   if matched_name != expected_name:
      logger.info(
         "Hoja '%s' resuelta como '%s' (score: %d).",
         expected_name,
         matched_name,
         score,
      )
   return matched_name
   

def verify_headers(ws: Worksheet, layout: type[LayoutA] | type[LayoutB], header_row: int, min_score: int = 85) -> bool:
   """
   Verifica que los headers de la hoja Excel correspondan con el layout esperado.
   Compara cada columna definida en el layout contra el valor real en la fila header_row.
   
   Returns:
       True  — todos los headers pasaron el umbral mínimo
       False — al menos un header falló (se loguean los detalles)
   """
   all_ok = True

   for member in layout:
      expected_label = member.label
      col            = member.col
      cell_value     = ws.cell(row=header_row, column=col).value

      if cell_value is None:
         logger.warning(
             "Header faltante en col %d: se esperaba '%s', celda vacía.",
             col, expected_label,
         )
         all_ok = False
         continue

      actual = str(cell_value).strip()
      score  = fuzz.WRatio(expected_label.upper(), actual.upper())

      if score < min_score:
         logger.warning(
             "Header col %d no coincide: esperado='%s', encontrado='%s', score=%d (mínimo=%d).",
             col, expected_label, actual, score, min_score,
         )
         all_ok = False
      else:
         if actual.upper() != expected_label.upper():
             logger.info(
                 "Header col %d resuelto: esperado='%s', encontrado='%s' (score=%d).",
                 col, expected_label, actual, score,
             )

   return all_ok

def depurate_lineup_files(folder_path : Path, offices_key : list[str], min_score : int = MIN_MATCH_SCORE):
   """
   Funcion para cargar los archivos de excel de las oficinas que esten en una ruta y develve las hojas esperadas para cada una 
   """

   xlsx_files = list(folder_path.glob('*.xlsx'))
   xlsx_files_names = [f.name.upper() for f in xlsx_files]

   if not xlsx_files:
      logger.warning("No se encontraron archivos .xlsx en '%s'.", folder_path)

   if len(xlsx_files) != len(offices_key):
      logger.warning(
         "No hay los archivos suficientes para procesar todas las oficinas"
      )

   # Creamos el checker
   checker = OverlapChecker()     
   # Para el registry habria que crearlo tambien desde aqui, e ir guardando los errores desde aca para cada officina y carga el libro (puede ser diccionario o algo asi)
   depuration_result = DepurationResult()
      
   for office in offices_key:
      result = process.extractOne(office,xlsx_files_names, scorer=fuzz.partial_ratio)
      if result is None:
         logger.warning(
            "No se encontró ninguna hoja para '%s'. Hojas disponibles: %s",
            office,
            xlsx_files_names,
         )
         continue
      office_match,score, idx = result
      if score < min_score:
         logger.warning(
            "Mejor coincidencia para '%s' fue '%s' con score %d (mínimo requerido: %d). "
            "Se retorna None.",
            office,
            office_match,
            score,
            min_score,
         )
         continue
      file_path = xlsx_files[idx]
      xlsx_files.pop(idx) ; xlsx_files_names.pop(idx)

      logger.info(f'Procesando archivo de la oficina %s con nombre %s ',office,file_path.name)
      file_expected_sheets = EXPECTED_SHEETS.get(office)
      assert file_expected_sheets is not None,f"No existe configuración de hojas para la oficina '{office}'. " f"Claves disponibles: {list(EXPECTED_SHEETS.keys())}"

      wb = openpyxl.load_workbook(file_path,data_only=True)

      available_sheets = wb.sheetnames
      logger.info(
         "Archivo '%s' cargado, Hojas disponibles: %s ",
         file_path.name,available_sheets
      )

      office_result = OfficeResult(office,file_path=file_path)
      
      for expected_name in file_expected_sheets:
         matched = find_best_sheet_match(available_sheets,expected_name)

         if matched is None:
            logger.warning("Hoja '%s' de oficina '%s' es None, se omite.", expected_name, office)
            continue

         available_sheets.remove(matched)
                  
         ws = wb[matched]
         
         layout = TERMINAL_LAYOUTS.get(expected_name, MAIN_LAYOUT )
         max_col = max(member.col for member in layout)
         
         # Antes de iterar, vamos a mirar si la primera fila (HEADER_ROW, col = 2), corresponde con buena concidencia > 90% al primer valor de layout
         headers_ok = verify_headers(ws,layout, HEADER_ROW)
         if not headers_ok:
            logger.warning(
               "Hoja '%s' de oficina '%s' tiene headers que no coinciden con layout '%s'. "
               "Se continúa procesando, pero los datos pueden ser incorrectos.",
               matched, office, layout.__name__,
             )
         registry = ErrorRegistry()
         sheet_rows = []
         for excel_row, row in enumerate(ws.iter_rows(min_row=HEADER_ROW+1,max_col=max_col,values_only=True), start=HEADER_ROW+1 ):
            if all(cell is None for cell in row):
               break
            
            row_dict = row_to_dict(row, layout)
            if layout == MAIN_LAYOUT:
               row_dict['WINDOWS'] = None

            try:
               validated = LineUpBaseModel.model_validate(row_dict)
               sheet_rows.append(validated.to_client_report())
               
               interval = build_interval(excel_row, validated,expected_name)
               if interval is not None:
                  checker.add(interval)               

            except PydanticValidationError as exc:
               for error in exc.errors():
                  loc = error["loc"]
                  
                  ctx = error.get("ctx", {})

                  # field validator → loc tiene el campo; model validator → loc vacío pero ctx tiene 'fields'
                  if loc:
                      fields_involved = [str(loc[0])]
                  else:
                      fields_involved = ctx.get("fields", ["MODEL"])

                  msg_template = error["msg"]

                  field_values = ctx.get("values", {})
                  try:
                      msg = msg_template.format(**field_values)
                      print(msg)
                  except (KeyError, ValueError):
                      msg = msg_template

                  for field_name in fields_involved:
                      col = field_to_col(field_name, layout)
                      val = field_values.get(field_name, error.get("input"))

                      logger.warning(
                          "Fila %d, col %d, campo '%s': %s",
                          excel_row, col, field_name, msg,
                      )
                      registry.add(
                        CellError(
                           row=excel_row, col=col, field=field_name,
                           msg=msg, value=val, source="pydantic"
                        )
                     )
         office_result.sheets.append(
            SheetResult(
               expected_name=expected_name,
               matched_name=matched,
               headers_ok=headers_ok,
               rows=sheet_rows,
               registry=registry,
            )
         )
      depuration_result.offices.append(office_result)
      wb.close()

   conflicts = checker.check()
   depuration_result.overlaps = conflicts
   for conflict in conflicts:
       logger.warning(
           "Solapamiento: VESSEL '%s' fila %d vs fila %d.",
           conflict.vessel, conflict.row_a, conflict.row_b
       )

   return depuration_result 

def row_to_dict( row : tuple, layout : type[LayoutA] | type[LayoutB])->dict[str, Any]:
   return {
      member.name: (row[member.col - 1] if member.col - 1 < len(row) else None)
      for member in layout
   }


def field_to_col(field_name: str, layout) -> int:
   """
   Mapea el nombre del campo Pydantic (member.name del enum, ej. DATE_OF_ARRIVAL)
   al número de columna Excel del layout. Retorna -1 si no se encuentra.
   """
   for member in layout:
      if member.name == field_name:
         return member.col
   return -1
